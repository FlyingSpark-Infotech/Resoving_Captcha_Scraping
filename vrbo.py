from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import sys, time
import openpyxl
from openpyxl.styles import Color, PatternFill, Border, Side, Font
import pandas, datetime
import ast, json
fill = PatternFill(
    start_color='FFFFFF',
    # start_color='808080',
    # end_color='FFFFFF',
    fill_type='solid'
)

box = Border(
    diagonal=Side(border_style="thin"),
)
box.diagonalUp=True

# basic_style = Style(font=Font(name='Microsoft YaHei')
#     , border=box
#     , fill=fill)

class VRBOHelper():

    def __init__(self):
        # self.url = "https://www.vrbo.com/1048405"
        #chrome_path = "/home/manoj/Downloads/chromedriver_linux64/chromedriver"
        self.driver = webdriver.Chrome("/xxx/xxxxx/xxxxxx/chromedriver_linux64/chromedriver") 
        self.driver.maximize_window()
        # self.driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        # self.driver.execute_script("scrollBy(0,250);")
        # self.driver.execute_script("window.scrollTo(0, 8500);")
        self.file_name = '/xxx/xxxxxx/xxxx/xxxxxxx/Copy of VRBO Data_rev(05-08-2019).xlsx'
    
    def getCalendarData(self, tables):
        data = []
        try:
            for table in tables:
                trs = table.find_elements_by_xpath(".//tr")
                # print("trs : ",len(trs))
                if len(trs):
                    for tr in trs:
                        tds = tr.find_elements_by_xpath(".//td")
                        # print("tds : ",len(tds))
                        if len(tds):
                            for td in tds:
                                day = None
                                rate = ""
                                half_rate = ""
                                is_half_rate = False
                                try:
                                    day = td.find_element_by_xpath(".//div[@class='day-template__day']").text
                                    half_rate = td.find_element_by_xpath(".//div/div").get_attribute("class")
                                    if half_rate == "day-template day-template--pm-unavailable-stay day-template--available-stay day-template--disabled " or half_rate == "day-template day-template--am-unavailable-stay day-template--available-stay ":
                                        is_half_rate = True
                                    rate = td.find_element_by_xpath(".//div[@class='day-template__rate']").text
                                except:
                                    pass

                                if is_half_rate:
                                    rate += "_Diagonal"
                                if day:
                                    data.append(rate)

                                # print("Day : ",day)
                                # print("Rate : ",rate)
        except:
            pass
        
        return data    
    
    def getItems(self):
        d = []
        items = []
        try:
            # new_row = ['data1', 'data2', 'data3', 'data4']

            wb = openpyxl.load_workbook(filename=self.file_name)
            ws = wb.get_sheet_by_name('Sheet1')
            # row = ws.max_row
            # row = ws.get_highest_row() + 1
            
            
            # for count, data in enumerate(ws['A'], 1):
            #     if isinstance(data.value, datetime.date):
            #         row = count

            row = 7
            urls = [x.value for x in list(ws['E'])[row:]]
            row += 1

            if len(urls):
                for url in urls:
                    items = []
                    print("row",row)
                    try:
                        self.driver.get(url)
                        self.driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
                        time.sleep(2)
                        self.driver.find_element_by_tag_name('body').send_keys(Keys.CONTROL + Keys.HOME)
                        # self.driver.execute_script("window.scrollTo(0, 2000);")
                        time.sleep(2)
                        property_number = str(url).split('vrbo.com/')[-1].split("/")[0].split("?")[0].strip()
                        print("property_number : ",property_number)
                        
                        location = self.driver.find_element_by_xpath("//div[@class='listing-overview__location micro']")
                        location_city = None
                        location_state = None 
                        try:
                            if location:
                                location = location.text
                                location_city = str(location).split(',')[0].strip()
                                location_state = str(location).split(',')[1].strip()
                            print("location_city : ",location_city)
                            print("location_state : ",location_state)
                        except:
                            pass

                        properties = self.driver.find_elements_by_xpath("//ul[@class='list-unstyled margin-bottom-0x']/li")
                        print("properties : ",len(properties))
                        house = None
                        sleeps = None
                        bedrooms = None
                        bathrooms = None
                        min_stay = None
                        for propery in properties:
                            house_field = propery.find_element_by_xpath(".//span[@class='listing-bullets__span']").text
                            if house_field and (str(house_field).lower() == "house" or str(house_field).lower() == "cabin"):
                                house_field = propery.find_elements_by_xpath(".//span[@class='listing-bullets__span']")[-1]
                                if house_field and "." in house_field.text:
                                    house = house_field.text.split()[0].strip()

                            sleeps_field = propery.find_element_by_xpath(".//span[@class='listing-bullets__span']").text
                            if sleeps_field and "sleep" in str(sleeps_field).lower():
                                sleeps = str(sleeps_field).split(":")[-1].strip()

                            bedrooms_field = propery.find_element_by_xpath(".//span[@class='listing-bullets__span']").text
                            if bedrooms_field and "bedroom" in str(bedrooms_field).lower():
                                bedrooms = str(bedrooms_field).split(":")[-1].strip()

                            bathroom_field = propery.find_element_by_xpath(".//span[@class='listing-bullets__span']").text
                            if bathroom_field and "bathroom" in str(bathroom_field).lower():
                                bathrooms = str(bathroom_field).split(":")[-1].strip()

                            half_bath = ''
                            half_bath_field = propery.find_element_by_xpath(".//span[@class='listing-bullets__span']").text
                            if half_bath_field and "half baths" in str(half_bath_field).lower():
                                half_bath = str(half_bath_field).split(":")[-1].strip()

                            min_stay_field = propery.find_element_by_xpath(".//span[@class='listing-bullets__span']").text
                            if min_stay_field and "min stay" in str(min_stay_field).lower():
                                min_stay = str(min_stay_field).split(":")[-1].replace("nights", "").replace("night", "").strip()
                            
                        print("house : ",house)
                        print("sleeps : ",sleeps)
                        print("bedrooms : ",bedrooms)
                        print("bathrooms : ",bathrooms)
                        print("min_stay : ",min_stay)

                        avg_per_night = self.driver.find_element_by_xpath("//meta[@property='og:price:amount']")
                        if avg_per_night:
                            avg_per_night = avg_per_night.get_attribute('content')
                        print("avg_per_night : ",avg_per_night)

                        try:
                            total_reviews = self.driver.find_element_by_xpath("//a[@class='reviews-summary__a']/strong")
                            if total_reviews:
                                total_reviews = str(total_reviews.text).split()[0].strip()
                        except:
                            total_reviews = ""
                        print("total_reviews : ",total_reviews)

                        try:
                            average_review = self.driver.find_element_by_xpath("//span[@class='review-summary__header-ratings-average']")
                            if average_review:
                                average_review = str(average_review.text).split("/")[0].strip()
                        except:
                            average_review = ""
                        print("average_review : ",average_review)
                        scripts = self.driver.find_elements_by_xpath("//script")
                        script = None
                        # for s in scripts:
                        #     if "window.__INITIAL_STATE__" in s.get_attribute("innerHTML"):
                        #         script = s.get_attribute("innerHTML")
        
                        #         break
                                
                        a=self.driver.find_elements_by_xpath("//div[@class='rate-details__fee-description']")
                        b=self.driver.find_elements_by_xpath("//div[@class='rate-details__fee-amount']")
                        cleaning_fee = ''
                        pet_fee = ''
                        security_deposit = ''
                        propery_damage_ins = ''
                        guest_fee = ''
                        for i,j in zip(a,b):
                            if "Cleaning Fee" in i.text:
                                cleaning_fee=j.text
                            elif "Additional Guest Fee" in i.text:
                                guest_fee=j.text
                            elif "Property Damage Insurance" in i.text:
                                propery_damage_ins=j.text
                            elif "Pet Fee" in i.text:
                                pet_fee=j.text
                            elif "Security Deposit" in i.text:
                                security_deposit=j.text


                        if script:
                            print("Working")
                            script_data = str(script).split("window.__INITIAL_STATE__ =")[-1]
                            script_data = script_data.split("window.__REQUEST_STATE__ =")[0].strip()[:-1]
                            script_data = json.loads(script_data)
                            if "flatFees" in script_data['listingReducer']['rateSummary']:
                                for fee in script_data['listingReducer']['rateSummary']['flatFees']:
                                    if fee['type'] == "CLEANING_FEE":
                                        cleaning_fee = fee["minAmount"]
                                        print("cleaning_fee",cleaning_fee)
                                    if fee["type"] == "PET_FEE":
                                        pet_fee = fee["minAmount"]
                                    if fee['type'] == "GUEST_FEE":
                                        guest_fee = fee["minAmount"]
                                    if fee["type"] == "CSA_PDP_FEE":
                                        propery_damage_ins = fee["minAmount"]
                                    if fee["type"] == "REFUNDABLE_DAMAGE_DEPOSIT":
                                        security_deposit = fee["minAmount"]
                                    
                        items.extend([
                            datetime.date.today().strftime("%B %d, %Y"), 
                            property_number, 
                            location_city, 
                            location_state, 
                            url,
                            house,
                            sleeps,
                            bedrooms,
                            bathrooms,
                            half_bath,
                            min_stay,
                            avg_per_night,
                            cleaning_fee,
                            pet_fee,
                            security_deposit,
                            propery_damage_ins,
                            guest_fee,
                            total_reviews,
                            average_review
                        ])
                        k = [''] * 139
                        items.extend(k)
                        
                        tables = self.driver.find_elements_by_xpath("//table[@class='month-table notranslate']")[:2]
                        print("tables : ",len(tables))
                        if len(tables):
                            tables_data = self.getCalendarData(tables)
                            if len(tables_data):
                                items.extend(tables_data)

                            print("before : ",len(items))                   
                            #try:
                            # next_month_btn = self.driver.find_element_by_xpath("//button/span[contains(text(), 'Next month')]")
                            # print(next_month_btn)
                            # if next_month_btn:
                            #     print("|"*10)
                            #     #next_month_btn.click()
                            #     self.driver.execute_script("arguments[0].click();", next_month_btn)
                            #     #webdriver.ActionChains(self.driver).move_to_element(next_month_btn).click(next_month_btn).perform()
                            #     time.sleep(3)
                            #     tables = []
                            #     tables_data = None
                            #     tables = self.driver.find_elements_by_xpath("//table[@class='month-table notranslate']")
                            #     print("tables : ",len(tables[:1]))
                            #     if len(tables):

                            #         tables_data = self.getCalendarData(tables[:1])
                            #         if len(tables_data):
                            #             items.extend(tables_data)
                            #except:
                                #pass       
                               
                            # print("after : ",len(items))
                        print("iiiiiiiiiiii : ")
                        print(items)
                        #print("####################",br_color)
                        d.append({
                            'row': row,
                            'items': items
                        })
                        # row += 1
                        # break
                    except:
                        print(sys.exc_info())
                        pass
                    row += 1

                    # if row >= 11:
                    #     break
                    print("rrrrrrrrrrrrrrrrrrrrrrrrrr rrrrrrrrrrrrrrrrrrrrrrr  : ",row)

                for k in d:
                    for col, entry in enumerate(k['items'], start=1):
                        # ws_cell = None
                        # box = None
                        # redFill = None
                        # ws_cell = ws.cell(row=k['row'], column=col)
                        # if col >= 20:
                        #     if entry:
                        #         # print("aaaaaaaaaaaaaaaa")
                        #         ws.cell(row=k['row'], column=col, value=entry).border = box
                        #         # ws_cell.border = box
                        #         # ws_cell.value=entry
                        #         # ws.insert_cols(ws_cell)
                        #     else:
                        #         # print("bbbbbbbbbbbbbb : ",ws_cell.value, ws_cell, len(entry))
                        #         # print(col, "empty", len(entry), ws_cell)
                        #         # box.diagonalUp=False
                                
                        #         # ws_cell.fill = fill
                        #         # ws_cell.value=entry
                        #         ws.cell(row=k['row'], column=col).fill = fill
                        # else:
                        ws.cell(row=k['row'], column=col, value=entry)
                    # print("uuuuuuuuuuu")
                    
                wb.save(self.file_name)
        except:
            print(sys.exc_info())
            pass

        # close webdriver session
        # self.driver.close()
        return d

    def cellfill(self):
        filepath="/xxxx/xxxxx/xxxxxx/xxxxxx/Copy of VRBO Data_rev(05-08-2019).xlsx"
        import openpyxl
        wb_obj = openpyxl.load_workbook(filepath) 
        sheet_obj = wb_obj.active
        max_col = sheet_obj.max_column 
        max_row=sheet_obj.max_row
        box = Border(
        diagonal=Side(border_style="thin", color='000000'),
        )
        redFill = PatternFill(start_color='808080',
        end_color='808080',
        fill_type='solid')
        box.diagonalUp=True
        for j in range(8,max_row + 1):
            for i in range(21, max_col + 1):

                cell_obj = sheet_obj.cell(row = j, column = i)

                try:
                    if "_Diagonal" in cell_obj.value and "$" in cell_obj.value  :

                        cell_obj.value = str(cell_obj.value).replace("_Diagonal", "").strip()
                        cell_obj.border = box
                    elif "_Diagonal" in cell_obj.value:
                        cell_obj.value = str(cell_obj.value).replace("_Diagonal", "").strip()
                        cell_obj.fill=redFill
                except:pass 
                if cell_obj.value is None:
                    cell_obj.fill=redFill

        wb_obj.save(filepath)


    def start(self):
        items = self.getItems()
        print("items : ",len(items))
        print('#'*100)
        




# main function
if __name__ == '__main__':
    # objVH is an instance for VRBOHelper.
    objVH = VRBOHelper()
    objVH.start()
    print("filling color to cell and insert diagonal in the cell ")
    objVH.cellfill()
